import bs4
import openpyxl
from openpyxl.styles import Border, Side
import datetime
import shutil
import os


# エクセルへの入力
def inputExcel(countStr, count, ws, s):
    if count == 1:
        print('file -> ' + s)
        a = 'A' + countStr
        ws[a].value = s
    elif count == 2:
        print('locked -> ' + s)
        b = 'B' + countStr
        ws[b].value = int(s)
    elif count == 3:
        print('perfectMatch -> ' + s)
        c = 'C' + countStr
        ws[c].value = int(s)
    elif count == 4:
        print('contextMatch -> ' + s)
        d = 'D' + countStr
        ws[d].value = int(s)
    elif count == 5:
        if type(s) == str:
            if s == "":
                s_zero = "0"
                print('crossfileRep -> ' + s_zero)
                e = 'E' + countStr
                ws[e].value = int(s_zero)
            else:
                print('crossfileRep -> ' + s)
                e = 'E' + countStr
                ws[e].value = int(s)
    elif count == 6:
        print('repeated -> ' + s)
        f = 'F' + countStr
        ws[f].value = int(s)
    elif count == 7:
        print('match100 -> ' + s)
        g = 'G' + countStr
        ws[g].value = int(s)
    elif count == 8:
        print('match99_95 -> ' + s)
        h = 'H' + countStr
        ws[h].value = int(s)
    elif count == 9:
        print('match94_85 -> ' + s)
        i = 'I' + countStr
        ws[i].value = int(s)
    elif count == 10:
        print('mat84_75 -> ' + s)
        j = 'J' + countStr
        ws[j].value = int(s)
    elif count == 11:
        print('match74_50 -> ' + s)
        k = 'K' + countStr
        ws[k].value = int(s)
    elif count == 12:
        print('new -> ' + s)
        l = 'L' + countStr
        ws[l].value = int(s)
    elif count == 13:
        print('totalCharacters -> ' + s)
        m = 'M' + countStr
        ws[m].value = int(s)


if __name__ == '__main__':
    s = input("File: ")
    log_file = s.strip('\"')
    #print(log_file)

    # パーサー
    soup = bs4.BeautifulSoup(open(log_file, 'r', encoding="utf-16"), 'html.parser')
    trs = soup.find_all('tr')

    # 結果が入るエクセルの準備
    xlsxTemplate = "log.xlsx"
    todaydetail = datetime.datetime.today()
    datetime = todaydetail.strftime("%Y%m%d%H%M%S")
    resultsFile = datetime + '_' + xlsxTemplate
    shutil.copyfile(xlsxTemplate, resultsFile)

    # Excelを準備
    wb = openpyxl.load_workbook(resultsFile)
    ws = wb['Characters']
    ws2 = wb['Words']

    # Main
    countT = 1
    flag = 0
    count = 1
    for tr in trs:
        for td in tr.find_all('td'):
            s = td.text
            
            # ワードカウント分が出てきたら新たにセットする
            if s == 'Total (words)':
                flag += 1
                countT = 2
                count = 1

            countStr = str(countT)

            if flag == 0:
                inputExcel(countStr, count, ws, s) # Charactorsシートに入力する
            elif flag <= 1:
                inputExcel(countStr, count, ws2, s) # Wordsシートに入力する

            if count <= 12:
                count += 1
            elif count == 13:
                count = 1

        countT += 1 # エクセルの入力行
    
    side = Side(style="thin", color="000000") # セルの線
    countT = countT - 1

    # セルの線を付ける
    for row in ws["A3":"M" + str(countT)]:
        for cell in row:
            cell.border = Border(left=side, right=side, top=side, bottom=side)
    
    for row in ws2["A3":"M" + str(countT)]:
        for cell in row:
            cell.border = Border(left=side, right=side, top=side, bottom=side)


    # Excelを閉じて保存
    wb.close()
    wb.save(resultsFile)

    print("\n" + 'Done!')
    os.system("pause > nul")
